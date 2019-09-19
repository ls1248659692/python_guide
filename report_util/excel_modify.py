# coding:utf-8
# !/usr/bin/env python

import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from  report_util.data_analysis_head import *

reload(sys)
sys.setdefaultencoding('utf-8')

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(PROJECT_ROOT, '../'))
db_url = 'mysql://admin:admin2048az@192.168.50.95:3306/ndb?charset=utf8'


class UpdateExcel(object):
    def __init__(self, filename, sheetname, header_row=1):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[sheetname]
        self.header_row = header_row
        self.value_row = header_row + 1
        self.value_col = 1

    def write_cell(self, cell_row, cell_column, value):
        assert isinstance(cell_row, int) and isinstance(
            cell_column, int), 'cell_row and cell_column must be an integer.'
        self.ws.cell(row=cell_row, column=cell_column).value = value

    def write_column(self, col, value_list=[]):
        assert isinstance(col, int), 'col must be an integer.'
        for row_num in range(
                self.value_row, min(
                            self.ws.max_row + 1, len(value_list) + self.value_row)):
            self.ws.cell(row=row_num,
                         column=col).value = value_list[row_num - self.value_row]

    def write_row(self, row, value_list=[]):
        assert isinstance(row, int), 'row must be an integer.'
        for col_num in range(
                self.value_col, min(
                            self.ws.max_column + 1, len(value_list) + self.value_col)):
            self.ws.cell(
                row=row, column=col_num).value = value_list[col_num - self.value_col]

    def write_df(self, df):
        self.wb.remove(self.ws)
        self.wb.create_sheet(self.sheetname)
        self.ws = self.wb[self.sheetname]
        dataset = [df.columns.tolist()] + df.values.tolist()

        for row in dataset:
            self.ws.append(row)

    def set_cell(self, cell_row, cell_column, font=None, align=None):
        assert isinstance(cell_row, int) and isinstance(
            cell_column, int), 'cell_row and cell_column must be an integer.'
        cell_info = self.ws.cell(row=cell_row, column=cell_column)
        cell_info.font = font
        cell_info.alignment = align

    def set_column(self, column_width={}, column_hidden=[]):
        if not column_width:
            for col in self.ws.columns:
                self.ws.column_dimensions[col[0].column].width = 20
        else:
            for col, col_width in column_width.items():
                col_str = get_column_letter(col) if isinstance(
                    col, (int, float)) else col
                self.ws.column_dimensions[col_str].width = col_width

        if column_hidden:
            for col in column_hidden:
                col_str = get_column_letter(col) if isinstance(
                    col, (int, float)) else col
                self.ws.column_dimensions.group(start=col_str, hidden=True)

    def set_format(self, col=1, is_header=False, font=None, align=None, number_format=None):
        assert isinstance(col, int), 'col must be an integer.'
        if is_header:
            for col_num in range(1, self.ws.max_column + 1):
                self.ws.cell(row=self.header_row, column=col_num).font = font
                self.ws.cell(
                    row=self.header_row,
                    column=col_num).alignment = align
        else:
            for row_num in range(2, self.ws.max_row + 1):
                self.ws.cell(row=row_num, column=col).font = font
                self.ws.cell(row=row_num, column=col).alignment = align
                self.ws.cell(row=row_num,
                             column=col).number_format = number_format

    def save(self):
        self.wb.save(self.filename)


def check_fdir():
    fdir = os.path.join(PROJECT_ROOT, 'revenue_report/')
    if not os.path.exists(fdir):
        os.mkdir(fdir)
    return fdir


def update_excel(filename, sheetname):
    ws = UpdateExcel(filename, sheetname)
    header_font = Font(name='calibri', size=14, bold=True)
    font = Font(name='calibri', size=12)
    align = Alignment(horizontal='center', vertical='center')
    date_fmt = "M/D/YYYY"

    df = pd.DataFrame(
        {
            'A': range(
                10,
                15),
            'B': range(5),
            'C': list('abcde'),
            'D': [True] * 5,
            'E': pd.date_range(
                '2015-07-03',
                '2015-07-07'),
            'I': [
                'dom',
                'psf',
                'ghk',
                'gus',
                'pe']})
    ws.write_df(df)
    ws.write_cell(2, 1, 'hello world!')
    ws.write_column(1, value_list=[1, 2, 3, 4])
    ws.write_column(3, value_list=['a', 'b', 'c', 'd'])

    ws.set_cell(1, 2, font=font, align=align)
    ws.set_column(column_width={5: 20}, column_hidden=[6])
    ws.set_format(is_header=True, font=header_font, align=align)
    ws.set_format(
        col=5,
        is_header=False,
        font=font,
        align=align,
        number_format=date_fmt)
    ws.save()


def main():
    fdir = check_fdir()
    update_excel(fdir + 'revenue_report_bysql.xlsx', 'Sheet1')


if __name__ == '__main__':
    main()
